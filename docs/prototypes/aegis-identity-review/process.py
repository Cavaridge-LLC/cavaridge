"""
AEGIS Identity Access Review — Prototype Processing Engine
===========================================================
Reference implementation of the IAR correlation and risk flag logic.
This is NOT production code. It serves as the executable specification
for the TypeScript rewrite in the Cavaridge monorepo.

Spec: CVG-AEGIS-IDENTITY-REVIEW-v1.0
Date: 2026-03-24

Inputs:
  1. Entra ID user export CSV (M365 Admin Center → Users → Export)
  2. M365 Active User Detail CSV (M365 Admin Center → Reports → Usage → Active Users)

Output:
  Branded XLSX with 5 tabs:
    - Executive Summary (metrics, recommendations)
    - All Users (full inventory with risk flags)
    - Flagged Users (filtered to accounts with risk flags)
    - External Guests (guest-specific analysis)
    - License Breakdown (licensed users with per-service detail)

Usage:
  python process.py <users_csv> <activity_csv> <output_xlsx> [--tenant-name "Client Name"]
"""

import argparse
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONSTANTS — Cavaridge Brand Standards
# =============================================================================

BLUE_HDR_HEX = '2E5090'
BAND_HEX = 'F2F6FA'
BORDER_HEX = 'BFBFBF'
FONT_NAME = 'Arial'

BLUE_HDR = PatternFill('solid', fgColor=BLUE_HDR_HEX)
LIGHT_BAND = PatternFill('solid', fgColor=BAND_HEX)
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
RED_FILL = PatternFill('solid', fgColor='FFF0F0')
AMBER_FILL = PatternFill('solid', fgColor='FFF3E0')

HDR_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name=FONT_NAME, size=9)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=9)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=BLUE_HDR_HEX)
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color='1A1A1A')
MUTED_FONT = Font(name=FONT_NAME, italic=True, size=10, color='555555')

THIN_BORDER = Border(
    left=Side(style='thin', color=BORDER_HEX),
    right=Side(style='thin', color=BORDER_HEX),
    top=Side(style='thin', color=BORDER_HEX),
    bottom=Side(style='thin', color=BORDER_HEX),
)


# =============================================================================
# DATA INGESTION
# =============================================================================

def load_and_normalize(users_path: str, activity_path: str) -> pd.DataFrame:
    """Load both CSVs, normalize UPNs, classify user types, and merge."""
    users = pd.read_csv(users_path, encoding='utf-8-sig')
    activity = pd.read_csv(activity_path, encoding='utf-8-sig')

    users['upn_lower'] = users['User principal name'].str.lower().str.strip()
    activity['upn_lower'] = activity['User Principal Name'].str.lower().str.strip()

    users['User Type'] = users['User principal name'].apply(classify_user_type)

    merged = users.merge(activity, on='upn_lower', how='left', suffixes=('', '_activity'))
    return merged, activity


def classify_user_type(upn: str) -> str:
    """Classify user as Member or External Guest based on UPN pattern."""
    if pd.isna(upn):
        return 'Unknown'
    return 'External Guest' if '#ext#' in str(upn).lower() else 'Member'


# =============================================================================
# DATE HANDLING
# =============================================================================

NOW = pd.Timestamp.now(tz='UTC')

def parse_date(val):
    if pd.isna(val) or str(val).strip() == '':
        return None
    try:
        return pd.to_datetime(val, utc=True)
    except Exception:
        return None


def days_since(val) -> int | None:
    d = parse_date(val)
    return (NOW - d).days if d is not None else None


# =============================================================================
# DERIVED FIELDS
# =============================================================================

ACTIVITY_DATE_COLS = [
    'Exchange Last Activity Date',
    'OneDrive Last Activity Date',
    'SharePoint Last Activity Date',
    'Skype For Business Last Activity Date',
    'Yammer Last Activity Date',
    'Teams Last Activity Date',
]


def compute_derived_fields(merged: pd.DataFrame, activity: pd.DataFrame) -> pd.DataFrame:
    """Add computed columns: blocked status, license status, activity dates, risk flags."""

    # Sign-in blocked
    merged['Sign-In Blocked'] = merged['Block credential'].apply(
        lambda x: 'Yes' if str(x).strip().lower() == 'true' else 'No'
    )

    # License status
    merged['License Status'] = merged['Licenses'].apply(
        lambda x: 'Unlicensed' if pd.isna(x) or str(x).strip().lower() in ('', 'unlicensed') else str(x).strip()
    )
    merged['Is Licensed'] = merged['License Status'].apply(lambda x: 'No' if x == 'Unlicensed' else 'Yes')

    # Deleted
    merged['Is Deleted'] = merged['Is Deleted'].apply(
        lambda x: 'Yes' if str(x).strip().lower() == 'true' else 'No'
    )

    # Most recent activity across all services
    def most_recent_activity(row):
        dates = [parse_date(row.get(col)) for col in ACTIVITY_DATE_COLS]
        dates = [d for d in dates if d is not None]
        return max(dates) if dates else None

    merged['Last Activity Date'] = merged.apply(most_recent_activity, axis=1)
    merged['Days Since Last Activity'] = merged['Last Activity Date'].apply(
        lambda d: (NOW - d).days if d else None
    )

    # Account age
    merged['Account Created'] = merged['When created'].apply(parse_date)
    merged['Account Age (Days)'] = merged['Account Created'].apply(
        lambda d: (NOW - d).days if d else None
    )

    # Risk flags
    activity_upns = set(activity['upn_lower'].values)
    merged['Risk Flags'] = merged.apply(lambda row: compute_risk_flags(row, activity_upns), axis=1)

    # Sort: flagged first, then by type, then by name
    merged['_sort_key'] = merged['Risk Flags'].apply(lambda x: 0 if x != 'None' else 1)
    merged = merged.sort_values(['_sort_key', 'User Type', 'Display name'])

    return merged


# =============================================================================
# RISK FLAG ENGINE
# =============================================================================

def compute_risk_flags(row: pd.Series, activity_upns: set) -> str:
    """
    Deterministic risk flag computation. No LLM required.
    Returns semicolon-delimited string of flags, or 'None'.

    Flag taxonomy (see spec §4.2):
      - Blocked but Licensed          (High)
      - External with License          (High)
      - Inactive Licensed >90d         (Medium)
      - Inactive Licensed >180d        (High — overrides >90d)
      - Licensed — No Activity Data    (Medium)
      - Stale External Guest           (Low)
    """
    flags = []

    is_blocked = row['Sign-In Blocked'] == 'Yes'
    is_licensed = row['Is Licensed'] == 'Yes'
    is_external = row['User Type'] == 'External Guest'
    days_inactive = row.get('Days Since Last Activity')
    account_age = row.get('Account Age (Days)')
    has_activity_data = row.get('upn_lower') in activity_upns

    if is_blocked and is_licensed:
        flags.append('Blocked but Licensed')

    if is_external and is_licensed:
        flags.append('External with License')

    if is_licensed:
        if days_inactive is not None and days_inactive > 90:
            flags.append(f'Inactive {int(days_inactive)}d (Licensed)')
        elif not has_activity_data:
            flags.append('Licensed — No Activity Data')

    if not is_blocked and not is_licensed and is_external:
        if account_age and account_age > 180:
            flags.append('Stale External Guest')

    return '; '.join(flags) if flags else 'None'


# =============================================================================
# XLSX FORMATTING HELPERS
# =============================================================================

def style_header_row(ws, row_num: int, max_col: int):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = BLUE_HDR
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row: int, end_row: int, max_col: int, risk_col: int | None = None):
    for r in range(start_row, end_row + 1):
        band = LIGHT_BAND if (r - start_row) % 2 == 0 else WHITE_FILL
        risk_val = ws.cell(row=r, column=risk_col).value if risk_col else None
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.fill = RED_FILL if (risk_val and risk_val != 'None') else band


def write_data_sheet(ws, df: pd.DataFrame, columns: list[str], col_widths: list[int],
                     risk_col_name: str | None = None, tab_color: str = '4472C4'):
    """Write a standard data sheet with headers, data, filters, and frozen panes."""
    ws.sheet_properties.tabColor = tab_color

    for c_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=c_idx, value=col_name)
    style_header_row(ws, 1, len(columns))

    risk_col_idx = (columns.index(risk_col_name) + 1) if risk_col_name else None

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        for c_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name)
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(val, pd.Timestamp):
                cell.value = val.strftime('%Y-%m-%d')
            elif pd.isna(val):
                cell.value = ''
            else:
                cell.value = val

    data_end = 1 + len(df)
    style_data_rows(ws, 2, data_end, len(columns), risk_col=risk_col_idx)
    ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}{data_end}'
    ws.freeze_panes = 'A2'

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =============================================================================
# REPORT GENERATION
# =============================================================================

def build_executive_summary(ws, merged: pd.DataFrame, tenant_name: str):
    """Write the Executive Summary tab."""
    ws.sheet_properties.tabColor = BLUE_HDR_HEX

    total = len(merged)
    members = len(merged[merged['User Type'] == 'Member'])
    externals = len(merged[merged['User Type'] == 'External Guest'])
    licensed = len(merged[merged['Is Licensed'] == 'Yes'])
    unlicensed = len(merged[merged['Is Licensed'] == 'No'])
    blocked = len(merged[merged['Sign-In Blocked'] == 'Yes'])
    blocked_licensed = len(merged[(merged['Sign-In Blocked'] == 'Yes') & (merged['Is Licensed'] == 'Yes')])
    ext_licensed = len(merged[(merged['User Type'] == 'External Guest') & (merged['Is Licensed'] == 'Yes')])
    flagged = len(merged[merged['Risk Flags'] != 'None'])
    inactive_90 = len(merged[
        (merged['Is Licensed'] == 'Yes') &
        (merged['Days Since Last Activity'].notna()) &
        (merged['Days Since Last Activity'] > 90)
    ])
    inactive_180 = len(merged[
        (merged['Is Licensed'] == 'Yes') &
        (merged['Days Since Last Activity'].notna()) &
        (merged['Days Since Last Activity'] > 180)
    ])
    no_activity = len(merged[
        (merged['Is Licensed'] == 'Yes') &
        (merged['Days Since Last Activity'].isna())
    ])

    ws['A1'] = f'{tenant_name} — Microsoft 365 User Security Analysis'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    ws['A2'] = f'Report Date: {NOW.strftime("%B %d, %Y")}'
    ws['A2'].font = SUBTITLE_FONT
    ws['A3'] = 'Prepared by: Dedicated IT (DIT)'
    ws['A3'].font = MUTED_FONT

    metrics = [
        ('Metric', 'Count', 'Notes'),
        ('Total User Objects', total, 'Includes members, guests, system/room accounts'),
        ('Internal Members', members, 'Accounts with internal UPN'),
        ('External Guests', externals, 'Accounts with #EXT# in UPN'),
        ('Licensed Users', licensed, 'Users with at least one M365 license assigned'),
        ('Unlicensed Users', unlicensed, 'No license assigned'),
        ('Sign-In Blocked', blocked, 'Credential block enabled'),
        ('Blocked but Still Licensed', blocked_licensed, '⚠ Immediate action — wasted licenses'),
        ('External Guests with Licenses', ext_licensed, '⚠ Review — external users should rarely hold licenses'),
        ('Licensed Users Inactive >90 Days', inactive_90, '⚠ Potential reclamation candidates'),
        ('Licensed Users Inactive >180 Days', inactive_180, '⚠ Strong reclamation candidates'),
        ('Licensed Users — No Activity Data', no_activity, 'No O365 activity match; verify manually'),
        ('Total Users with Risk Flags', flagged, 'Users appearing on Flagged Users tab'),
    ]

    start = 5
    for i, (m, c, n) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=c)
        ws.cell(row=r, column=3, value=n)

    style_header_row(ws, start, 3)
    for r in range(start + 1, start + len(metrics)):
        band = LIGHT_BAND if (r - start) % 2 == 0 else WHITE_FILL
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = band
            if c == 1:
                cell.font = BOLD_FONT
        notes = ws.cell(row=r, column=3).value
        if notes and '⚠' in str(notes):
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = AMBER_FILL

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 65

    # Recommendations
    rec_r = start + len(metrics) + 2
    ws.cell(row=rec_r, column=1, value='Key Recommendations').font = SUBTITLE_FONT
    recommendations = [
        'Remove licenses from all blocked accounts to reclaim spend immediately.',
        'Review all external guest accounts with assigned licenses — revoke unless business-justified.',
        'Disable or delete licensed accounts inactive for 180+ days after stakeholder confirmation.',
        'Investigate licensed accounts with no activity data — may be service accounts or misconfigured.',
        'Implement a recurring quarterly user access review cadence.',
        'Purge stale external guest accounts (>180 days, unlicensed, no activity).',
    ]
    for i, rec in enumerate(recommendations):
        ws.cell(row=rec_r + 1 + i, column=1, value=f'{i + 1}. {rec}').font = BODY_FONT
        ws.merge_cells(start_row=rec_r + 1 + i, start_column=1, end_row=rec_r + 1 + i, end_column=3)


def generate_report(merged: pd.DataFrame, output_path: str, tenant_name: str):
    """Generate the full 5-tab XLSX report."""
    wb = Workbook()

    # Tab 1: Executive Summary
    ws_sum = wb.active
    ws_sum.title = 'Executive Summary'
    build_executive_summary(ws_sum, merged, tenant_name)

    # Tab 2: All Users
    ws_all = wb.create_sheet('All Users')
    write_data_sheet(ws_all, merged,
        columns=['Display name', 'User Type', 'Sign-In Blocked', 'Is Licensed', 'License Status',
                 'Title', 'Department', 'Account Created', 'Last Activity Date', 'Days Since Last Activity',
                 'Has Exchange License', 'Has Teams License', 'Has OneDrive License', 'Has SharePoint License',
                 'Risk Flags', 'User principal name'],
        col_widths=[30, 15, 16, 12, 35, 25, 22, 14, 14, 12, 10, 10, 10, 10, 40, 45],
        risk_col_name='Risk Flags', tab_color='4472C4')

    # Tab 3: Flagged Users
    ws_flag = wb.create_sheet('Flagged Users')
    write_data_sheet(ws_flag, merged[merged['Risk Flags'] != 'None'],
        columns=['Display name', 'User Type', 'Risk Flags', 'Sign-In Blocked', 'Is Licensed',
                 'License Status', 'Last Activity Date', 'Days Since Last Activity', 'User principal name'],
        col_widths=[30, 15, 45, 16, 12, 35, 14, 12, 45],
        risk_col_name='Risk Flags', tab_color='FF4444')

    # Tab 4: External Guests
    ws_ext = wb.create_sheet('External Guests')
    write_data_sheet(ws_ext, merged[merged['User Type'] == 'External Guest'],
        columns=['Display name', 'Sign-In Blocked', 'Is Licensed', 'License Status',
                 'Account Created', 'Account Age (Days)', 'Last Activity Date', 'Days Since Last Activity',
                 'Risk Flags', 'User principal name'],
        col_widths=[30, 16, 12, 35, 14, 14, 14, 12, 40, 50],
        risk_col_name='Risk Flags', tab_color='FFA500')

    # Tab 5: License Breakdown
    ws_lic = wb.create_sheet('License Breakdown')
    write_data_sheet(ws_lic, merged[merged['Is Licensed'] == 'Yes'],
        columns=['Display name', 'User Type', 'Sign-In Blocked', 'License Status',
                 'Has Exchange License', 'Has OneDrive License', 'Has SharePoint License',
                 'Has Teams License', 'Has Yammer License', 'Last Activity Date',
                 'Days Since Last Activity', 'Assigned Products'],
        col_widths=[30, 15, 16, 40, 10, 10, 10, 10, 10, 14, 12, 50],
        tab_color='00B050')

    wb.save(output_path)
    return wb


# =============================================================================
# CLI ENTRY POINT
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='AEGIS Identity Access Review — Report Generator')
    parser.add_argument('users_csv', help='Path to Entra ID user export CSV')
    parser.add_argument('activity_csv', help='Path to M365 Active User Detail CSV')
    parser.add_argument('output_xlsx', help='Path for output XLSX report')
    parser.add_argument('--tenant-name', default='Microsoft 365 Tenant',
                        help='Client/tenant name for report header')
    args = parser.parse_args()

    print(f'Loading data...')
    merged, activity = load_and_normalize(args.users_csv, args.activity_csv)

    print(f'Computing derived fields and risk flags...')
    merged = compute_derived_fields(merged, activity)

    print(f'Generating report...')
    generate_report(merged, args.output_xlsx, args.tenant_name)

    # Summary
    total = len(merged)
    flagged = len(merged[merged['Risk Flags'] != 'None'])
    licensed = len(merged[merged['Is Licensed'] == 'Yes'])
    print(f'\nComplete: {args.output_xlsx}')
    print(f'  Users: {total} | Licensed: {licensed} | Flagged: {flagged}')


if __name__ == '__main__':
    main()
